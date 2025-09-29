import sys, json, argparse, random, re, os, shutil
sys.path.append("src/")
import numpy as np
import pandas as pd
import logging
from datetime import datetime
from pathlib import Path
import math
import os.path as osp
import networkx as nx
import pdb
import pickle
import csv
import pandas as pd
from openpyxl import load_workbook
import os

import torch
import torch.nn as nn
import torch.nn.functional as func
from torch import optim
import torch.multiprocessing as mp
from torch_geometric.data import Data, Batch, DataLoader
from torch_geometric.utils import to_dense_batch, k_hop_subgraph

from utils import common_tools as ct
from utils.my_math import masked_mae_np, masked_mape_np, masked_mse_np
from utils.data_convert import generate_samples
from src.model.modeld import Basic_Model
from model.ewcc import EWC
from src.Dataset import COINDataset
from src.model import detect
from src.model import replay
from utils.DMC import DM


result = {3:{"mae":{}, "mape":{}, "rmse":{}}, 6:{"mae":{}, "mape":{}, "rmse":{}}, 12:{"mae":{}, "mape":{}, "rmse":{}}}
result_new = {3:{"mae":{}, "mape":{}, "rmse":{}}, 6:{"mae":{}, "mape":{}, "rmse":{}}, 12:{"mae":{}, "mape":{}, "rmse":{}}}
result_old = {3:{"mae":{}, "mape":{}, "rmse":{}}, 6:{"mae":{}, "mape":{}, "rmse":{}}, 12:{"mae":{}, "mape":{}, "rmse":{}}}
pin_memory = False
n_work = 0 

def update(src, tmp):
    for key in tmp:
        if key!= "gpuid":
            src[key] = tmp[key]


def load_best_model(args):
    if (args.load_first_year and args.year <= args.begin_year+1) or args.train == 0:
        load_path = args.first_year_model_path
        loss = load_path.split("/")[-1].replace(".pkl", "")
    else:
        loss = []
        for filename in os.listdir(osp.join(args.model_path, args.logname+args.time, str(args.year-1))): 
            loss.append(filename[0:-4])
        loss = sorted(loss)
        load_path = osp.join(args.model_path, args.logname+args.time, str(args.year-1), loss[0]+".pkl")

    args.logger.info("[*] load from {}".format(load_path))
    state_dict = torch.load(load_path, map_location=args.device)["model_state_dict"]
    if 'tcn2.weight' in state_dict:
        del state_dict['tcn2.weight']
        del state_dict['tcn2.bias']
    model = Basic_Model(args)
    model.load_state_dict(state_dict)
    model = model.to(args.device)
    return model

def init(args):    
    conf_path = osp.join(args.conf)
    info = ct.load_json_file(conf_path)
    info["time"] = datetime.now().strftime("%Y-%m-%d-%H:%M:%S.%f")
    update(vars(args), info)
    vars(args)["path"] = osp.join(args.model_path, args.logname+args.time)
    ct.mkdirs(args.path)
    del info


def init_log(args):
    log_dir, log_filename = args.path, args.logname
    logger = logging.getLogger(__name__)
    ct.mkdirs(log_dir)
    logger.setLevel(logging.INFO)
    fh = logging.FileHandler(osp.join(log_dir, log_filename+".log"))
    fh.setLevel(logging.INFO)
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    formatter = logging.Formatter("%(asctime)s - %(message)s")
    fh.setFormatter(formatter)
    ch.setFormatter(formatter)
    logger.addHandler(fh)
    logger.addHandler(ch)
    logger.info("logger name:%s", osp.join(log_dir, log_filename+".log"))
    vars(args)["logger"] = logger
    return logger


def seed_set(seed=0):
    max_seed = (1 << 32) - 1
    random.seed(seed)
    np.random.seed(random.randint(0, max_seed))
    torch.manual_seed(random.randint(0, max_seed))
    torch.cuda.manual_seed(random.randint(0, max_seed))
    torch.cuda.manual_seed_all(random.randint(0, max_seed))
    torch.backends.cudnn.benchmark = False  # if benchmark=True, deterministic will be False
    torch.backends.cudnn.deterministic = True

def neighborG(adj):
    neighborL = []
    for i in range(len(adj)):
        neighbor = []
        for j in range(len(adj)):
            if adj[i][j] != 0:
                neighbor.append(j)
        neighborL.append(neighbor)
    return neighborL

def train(inputs, inputsP,inputsF, args):
    # Model Setting
    global result
    global result_new
    global result_old
    path = osp.join(args.path, str(args.year))
    ct.mkdirs(path)
    if args.loss == "mse": lossfunc = func.mse_loss
    elif args.loss == "huber": lossfunc = func.smooth_l1_loss

    # Dataset Definition
    if args.strategy == 'incremental' and args.year > args.begin_year:
        graph = nx.Graph()
        graph.add_nodes_from(range(args.subgraph.size(0)))
        graph.add_edges_from(args.subgraph_edge_index.numpy().T)
        adj = nx.to_numpy_array(graph)
        adj = adj / (np.sum(adj, 1, keepdims=True) + 1e-6)
        vars(args)["sub_adj"] = torch.from_numpy(adj).to(torch.float).to(args.device)
        
        train_x,train_y,val_x,val_y = inputs['train_x'],inputs['train_y'],inputs['val_x'][:,:,args.subgraph.numpy()],inputs['val_y'][:,:,args.subgraph.numpy()]

        PN = args.sub_adj.size(0)
        sliceidx = args.shell.currentbuffer.idx
        sidx = list()
        for i in range(sliceidx.size(0)):
            sidx.append(int(sliceidx[i].item()))
        sidx = np.array(sidx)
        examples1=torch.tensor(train_x[sidx,:,:]).transpose(1,2).to(args.device)
        labels1 = torch.tensor(train_y[sidx,:,:]).transpose(1,2).to(args.device)
        year1 = torch.ones(args.buffer_size,len(args.subgraph.numpy()))*args.year
        nodeID = torch.Tensor(np.tile(args.subgraph.numpy(), (args.buffer_size, 1)))
        labels1 = torch.Tensor(labels1).to(args.device)
        year1 = year1.to(args.device)
        nodeID = nodeID.to(args.device)
        sliceidx = sliceidx.to(args.device)
        args.shell.buffer.empty()
        args.time_stamp.append(np.asarray(sliceidx.cpu()))
        args.shell.buffer.add_data(examples=examples1,
                             labels=labels1,
                             nodeID = nodeID,
                             year = year1,
                             idx = sliceidx)
        args.shell.currentbuffer.empty()

        if args.current_buffer_initiate:
            args.shell.currentbuffer.add_data(examples=examples1,
                             labels=labels1,
                             nodeID = nodeID,
                             year = year1,
                             idx = sliceidx,
                             score = torch.zeros(args.buffer_size))

        #idx = np.array(range(5356))
        idx = np.array(range(len(inputs['train_x'])))
        train_loader = DataLoader(COINDataset("", "", args.subgraph.numpy(),args.year, x=inputs['train_x'], y=inputs['train_y'], \
            idx = idx, edge_index="", mode="subgraph"), batch_size=args.batch_size, shuffle=True, pin_memory=pin_memory, num_workers=n_work)
        val_loader = DataLoader(COINDataset("", "", args.subgraph.numpy(),args.year, x=val_x, y=val_y, \
            idx = idx, edge_index="", mode="subgraph"), batch_size=args.batch_size, shuffle=False, pin_memory=pin_memory, num_workers=n_work)
        test_loader = DataLoader(COINDataset(inputsF, "test",np.array(range(args.adjf.size(0))),args.year+1,idx = idx), batch_size=args.batch_size, shuffle=False, pin_memory=pin_memory, num_workers=n_work)
    else:
        #idx = np.array(range(5356))
        idx = np.array(range(len(inputs['train_x'])))
        vars(args)['nodeID'] = np.array(range(args.adj.size(0)))
        vars(args)['fnodeID'] = np.array(range(args.adjf.size(0)))
        vars(args)["sub_adj"] = vars(args)["adj"]
        train_loader = DataLoader(COINDataset(inputs, "train",args.nodeID,args.year,idx = idx), batch_size=args.batch_size, shuffle=True, pin_memory=pin_memory, num_workers=n_work)
        val_loader = DataLoader(COINDataset(inputs,"val",args.nodeID,args.year,idx = idx), batch_size=args.batch_size, shuffle=False, pin_memory=pin_memory, num_workers=n_work)
        test_loader = DataLoader(COINDataset(inputsF, "test",args.fnodeID,args.year+1,idx = idx), batch_size=args.batch_size, shuffle=False, pin_memory=pin_memory, num_workers=n_work)

    
    args.logger.info("[*] Year " + str(args.year) + " Dataset load!")

    # Model Definition
    if args.init == True and args.year > args.begin_year:
        gnn_model= load_best_model(args) 
        if args.ewc:
            vars(args)['buf_data'] = examples1
            args.logger.info("[*] EWC! lambda {:.6f}".format(args.ewc_lambda))
            model = EWC(args, gnn_model,args.sub_adj, args.adj, args.adjp, args.subgraph.numpy(),args.cneighbors, sliceidx, args.warn_list, args.ewc_lambda, args.ewc_strategy)
            ewc_loaderc = DataLoader(COINDataset(inputs, "train",args.nodeID,args.year,idx = idx), batch_size=args.batch_size, shuffle=False, pin_memory=pin_memory, num_workers=n_work)
            ewc_loaderp = DataLoader(COINDataset(inputsP, "train",args.nodeID,args.year,idx = idx), batch_size=args.batch_size, shuffle=False, pin_memory=pin_memory, num_workers=n_work)
            model.register_ewc_params(ewc_loaderp,ewc_loaderc, lossfunc, device)
        else:
            model = gnn_model
    else:
        gnn_model = Basic_Model(args).to(args.device)
        model = gnn_model

    # Model Optimizer
    vars(args)["optimizer"]= torch.optim.Adam(model.parameters(), lr=args.lr)

    args.logger.info("[*] Year " + str(args.year) + " Training start")

    lowest_validation_loss = 1e7
    model.train()
    use_time = []
    N = args.sub_adj.size(0)
    M1 = args.batch_size * 2
    M2 = (len(inputs['train_x']) % args.batch_size )* 2
    M3 = args.validation_size
    Fbatch1 = torch.empty(M1, N, dtype = torch.int64)
    for i in range(N):
        Fbatch1[:, i] = torch.arange(0, M1, dtype = torch.int64)
    Fbatch1 = Fbatch1.reshape((M1*N))
    Fbatch2 = torch.empty(M2, N, dtype = torch.int64)
    for i in range(N):
        Fbatch2[:, i] = torch.arange(0, M2, dtype = torch.int64)
    Fbatch2 = Fbatch2.reshape((M2*N))
    Fbatch3 = torch.empty(M3, N, dtype = torch.int64)
    for i in range(N):
        Fbatch3[:, i] = torch.arange(0, M3, dtype = torch.int64)
    Fbatch3 = Fbatch3.reshape((M3*N))
    vars(args)['Fbatch1'] = Fbatch1.to(device)
    vars(args)['Fbatch2'] = Fbatch2.to(device)
    vars(args)['Fbatch3'] = Fbatch3.to(device)
    vars(args)['begin_influence'] = False
    for epoch in range(args.epoch):
        vars(args)["Nepoch"]=epoch
        training_loss = 0.0
        start_time = datetime.now()
        
        # Train Model
        cn = 0
        for batch_idx, data in enumerate(train_loader):
            if epoch == 0 and batch_idx == 0:
                args.logger.info("node number {}".format(data.x.shape))
            data = data.to(device, non_blocking=pin_memory)
            loss = args.shell.observe(data, args, model,lossfunc)
            training_loss += float(loss)
            loss.backward()
            args.optimizer.step()
            
            cn += 1

        if epoch == 0:
            total_time = (datetime.now() - start_time).total_seconds()
        else:
            total_time += (datetime.now() - start_time).total_seconds()
        use_time.append((datetime.now() - start_time).total_seconds())
        training_loss = training_loss/cn 
 
        # Validate Model
        validation_loss = 0.0
        cn = 0
        with torch.no_grad():
            for batch_idx, data in enumerate(val_loader):
                data = data.to(device,non_blocking=pin_memory)
                loss = args.shell.observeV(data, args, model,lossfunc)
                validation_loss += float(loss)
                cn += 1
        validation_loss = float(validation_loss/cn)
        

        args.logger.info(f"epoch:{epoch}, training loss:{training_loss:.4f} validation loss:{validation_loss:.4f}")

        # Early Stop
        """
        if validation_loss <= lowest_validation_loss:
            counter = 0
            lowest_validation_loss = round(validation_loss, 4)
            torch.save({'model_state_dict': gnn_model.state_dict()}, osp.join(path, str(round(validation_loss,4))+".pkl"))
        else:
            counter += 1
            if counter > patience:
                args.begin_influence = True
        if args.begin_influence:
            select_count = select_count + 1
        if select_count == args.start_selection:
            break
        """
        
        
        if validation_loss <= lowest_validation_loss:
            lowest_validation_loss = round(validation_loss, 4)
            torch.save({'model_state_dict': gnn_model.state_dict()}, osp.join(path, str(round(validation_loss,4))+".pkl"))

    best_model_path = osp.join(path, str(lowest_validation_loss)+".pkl")
    best_model = Basic_Model(args)
    best_model.load_state_dict(torch.load(best_model_path, args.device)["model_state_dict"])
    best_model = best_model.to(args.device)
    
    # Test Model
    test_model(best_model, args, test_loader, pin_memory, lossfunc)
    result[args.year] = {"total_time": total_time, "average_time": sum(use_time)/len(use_time), "epoch_num": epoch+1}
    args.logger.info("Finished optimization, total time:{:.2f} s, best model:{}".format(total_time, best_model_path))


def test_model(model, args, testset, pin_memory, lossfunc):
    model.eval()
    lossfunc = args.loss
    pred_ = []
    truth_ = []
    loss = 0.0
    with torch.no_grad():
        cn = 0
        for data in testset:
            data = data.to(args.device, non_blocking=pin_memory)
            Loss, pred = args.shell.observeT(data, args, model,lossfunc)
            loss += Loss
            pred, _ = to_dense_batch(pred, batch=data.batch)
            data.y, _ = to_dense_batch(data.y, batch=data.batch)
            pred_.append(pred.cpu().data.numpy())
            truth_.append(data.y.cpu().data.numpy())
            cn += 1
        loss = loss/cn
        args.logger.info("[*] loss:{:.4f}".format(loss))
        pred_ = np.concatenate(pred_, 0)
        truth_ = np.concatenate(truth_, 0)
        mae = metric(truth_, pred_, args)
        pred_new ,truth_new= pred_[:,args.en[args.year-args.begin_year]:,:], truth_[:,args.en[args.year-args.begin_year]:,:]
        mae = metric_new(truth_new, pred_new, args)
        pred_old ,truth_old= pred_[:,:args.en[args.year-args.begin_year],:], truth_[:,:args.en[args.year-args.begin_year],:]
        mae = metric_old(truth_old, pred_old, args)
        return loss


def metric(ground_truth, prediction, args):
    global result
    pred_time = [3,6,12]
    args.logger.info("[*] year {}, testing".format(args.year))
    for i in pred_time:
        mae = masked_mae_np(ground_truth[:, :, :i], prediction[:, :, :i], 0)
        rmse = masked_mse_np(ground_truth[:, :, :i], prediction[:, :, :i], 0) ** 0.5
        mape = masked_mape_np(ground_truth[:, :, :i], prediction[:, :, :i], 0)
        args.logger.info("T:{:d}\tMAE\t{:.4f}\tRMSE\t{:.4f}\tMAPE\t{:.4f}".format(i,mae,rmse,mape))
        result[i]["mae"][args.year] = mae
        result[i]["mape"][args.year] = mape
        result[i]["rmse"][args.year] = rmse
    return mae

def metric_new(ground_truth, prediction, args):
    global result_new
    pred_time = [3,6,12]
    args.logger.info("[*] year {}, testing".format(args.year))
    for i in pred_time:
        mae = masked_mae_np(ground_truth[:, :, :i], prediction[:, :, :i], 0)
        rmse = masked_mse_np(ground_truth[:, :, :i], prediction[:, :, :i], 0) ** 0.5
        mape = masked_mape_np(ground_truth[:, :, :i], prediction[:, :, :i], 0)
        args.logger.info("T:{:d}\tMAE\t{:.4f}\tRMSE\t{:.4f}\tMAPE\t{:.4f}".format(i,mae,rmse,mape))
        result_new[i]["mae"][args.year] = mae
        result_new[i]["mape"][args.year] = mape
        result_new[i]["rmse"][args.year] = rmse
    return mae

def metric_old(ground_truth, prediction, args):
    global result_old
    pred_time = [3,6,12]
    args.logger.info("[*] year {}, testing".format(args.year))
    for i in pred_time:
        mae = masked_mae_np(ground_truth[:, :, :i], prediction[:, :, :i], 0)
        rmse = masked_mse_np(ground_truth[:, :, :i], prediction[:, :, :i], 0) ** 0.5
        mape = masked_mape_np(ground_truth[:, :, :i], prediction[:, :, :i], 0)
        args.logger.info("T:{:d}\tMAE\t{:.4f}\tRMSE\t{:.4f}\tMAPE\t{:.4f}".format(i,mae,rmse,mape))
        result_old[i]["mae"][args.year] = mae
        result_old[i]["mape"][args.year] = mape
        result_old[i]["rmse"][args.year] = rmse
    return mae


def main(args):
    logger = init_log(args)
    logger.info("params : %s", vars(args))
    ct.mkdirs(args.save_data_path)
    vars(args)["shell"] = DM(args) #Creating memory buffer


    for year in range(args.begin_year, args.end_year):
        
        vars(args)['year'] = year

        # Load current year data
        inputs = generate_samples(31, osp.join(args.save_data_path, str(year)+'_30day'), np.load(osp.join(args.raw_data_path, str(year)+".npz"))["x"], graph, val_test_mix=True) \
            if args.data_process else np.load(osp.join(args.save_data_path, str(year)+".npz"), allow_pickle=True)
        graph = nx.from_numpy_array(np.load(osp.join(args.graph_path, str(year)+"_adj.npz"))["x"])
        vars(args)["graph_size"] = graph.number_of_nodes()
        args.logger.info("[*] Year {} load from {}.npz".format(year, osp.join(args.save_data_path, str(year)))) 

        adj = np.load(osp.join(args.graph_path, str(args.year)+"_adj.npz"))["x"]
        adj = adj / (np.sum(adj, 1, keepdims=True) + 1e-6)
        vars(args)["adj"] = torch.from_numpy(adj).to(torch.float).to(args.device)

        # Load next year data
        inputsF = generate_samples(31, osp.join(args.save_data_path, str(year+1)+'_30day'), np.load(osp.join(args.raw_data_path, str(year+1)+".npz"))["x"], graph, val_test_mix=True) \
            if args.data_process else np.load(osp.join(args.save_data_path, str(year+1)+".npz"), allow_pickle=True)
        fgraph = nx.from_numpy_array(np.load(osp.join(args.graph_path, str(year+1)+"_adj.npz"))["x"])
        vars(args)["fgraph_size"] = fgraph.number_of_nodes()        
        args.logger.info("[*] Year {} load from {}.npz".format(year+1, osp.join(args.save_data_path, str(year)))) 

        adjf = np.load(osp.join(args.graph_path, str(args.year+1)+"_adj.npz"))["x"]
        adjf = adjf / (np.sum(adjf, 1, keepdims=True) + 1e-6)
        vars(args)["adjf"] = torch.from_numpy(adjf).to(torch.float).to(args.device)
        inputsP = inputs # If the year is the begin year, then inputsP = inputs


        if year == args.begin_year and args.load_first_year:
            # Skip the first year, model has been trained and retrain is not needed
            model= load_best_model(args)
            test_loader = DataLoader(COINDataset(inputs, "test"), batch_size=args.batch_size, shuffle=False, pin_memory=pin_memory, num_workers=n_work)
            test_model(model, args, test_loader, pin_memory=True)
            continue

        if year > args.begin_year and args.strategy == "incremental":
         # Load previous year data
            inputsP = generate_samples(31, osp.join(args.save_data_path, str(year-1)+'_30day'), np.load(osp.join(args.raw_data_path, str(year-1)+".npz"))["x"], graph, val_test_mix=True) \
            if args.data_process else np.load(osp.join(args.save_data_path, str(year-1)+".npz"), allow_pickle=True)
        
            args.logger.info("[*] Year {} load from {}.npz".format(year-1, osp.join(args.save_data_path, str(year)))) 

            adjp = np.load(osp.join(args.graph_path, str(args.year-1)+"_adj.npz"))["x"]
            adjp = adjp / (np.sum(adjp, 1, keepdims=True) + 1e-6)
            vars(args)["adjp"] = torch.from_numpy(adjp).to(torch.float).to(args.device)
            # Load the best model
            #model= load_best_model(args)
            vars(args)["pg"] = np.array(list(nx.from_numpy_array(np.load(osp.join(args.graph_path, str(year-1)+"_adj.npz"))["x"]).edges)).T
            vars(args)["cg"] = np.array(list(nx.from_numpy_array(np.load(osp.join(args.graph_path, str(year)+"_adj.npz"))["x"]).edges)).T           

            vars(args)["neighbors"] = neighborG(adjp)
            vars(args)["cneighbors"] = neighborG(adj)
            file_path = str(year)+'n.pkl'
            with open(file_path, 'wb') as file:
                pickle.dump(args.cneighbors, file)
            vars(args)["fneighbors"] = neighborG(adjf)
            
            model= load_best_model(args)
            node_list = list()
            # Obtain increase nodes
            if args.increase:
                pre_node_size = np.load(osp.join(args.graph_path, str(year-1)+"_adj.npz"))["x"].shape[0]
                cur_node_size = np.load(osp.join(args.graph_path, str(year)+"_adj.npz"))["x"].shape[0]
                node_list.extend(list(range(pre_node_size, cur_node_size)))



            # Obtain influence nodes
            if args.detect:
                args.logger.info("[*] detect strategy {}".format(args.detect_strategy))
                if args.type == 'traffic':
                    pre_data = np.load(osp.join(args.raw_data_path, str(year-1)+".npz"))['x']
                    cur_data = np.load(osp.join(args.raw_data_path, str(year)+".npz"))['x']
                else:
                    pre_data = np.load(osp.join(args.raw_data_path, str(year-1)+".npz"))['train_x']
                    cur_data = np.load(osp.join(args.raw_data_path, str(year)+".npz"))['train_x']
                pre_graph = np.array(list(nx.from_numpy_array(np.load(osp.join(args.graph_path, str(year-1)+"_adj.npz"))["x"]).edges)).T
                cur_graph = np.array(list(nx.from_numpy_array(np.load(osp.join(args.graph_path, str(year)+"_adj.npz"))["x"]).edges)).T
                # 20% of current graph size will be sampled
                vars(args)["topk"] = int(0.01*args.graph_size)
                vars(args)["topkR"] = int(args.warn_ratio*args.graph_size)
                influence_node_list_pro, warn_list = detect.influence_node_selection_promax(model, args, pre_data, cur_data, pre_graph, cur_graph)
                node_list.extend(list(influence_node_list_pro))
                node_list.extend(list(warn_list))

 


            # Obtain sample nodes
            if args.replay:
                vars(args)["replay_num_samples"] = int(0.09*args.graph_size) #int(0.2*args.graph_size)- len(node_list)
                args.logger.info("[*] replay node number {}".format(args.replay_num_samples))
                replay_node_list = replay.replay_node_selection(args, inputs, model)
                node_list.extend(list(replay_node_list))
            
            node_list = list(set(node_list))
            if len(node_list) > int(args.inform_size*args.graph_size):
                node_list = random.sample(node_list, int(0.1*args.graph_size))
            
            # Obtain subgraph of new node list
            cur_graph = torch.LongTensor(np.array(list(nx.from_numpy_array(np.load(osp.join(args.graph_path, str(year)+"_adj.npz"))["x"]).edges)).T)
            edge_list = list(nx.from_numpy_array(np.load(osp.join(args.graph_path, str(year)+"_adj.npz"))["x"]).edges)
            graph_node_from_edge = set()
            for (u,v) in edge_list:
                graph_node_from_edge.add(u)
                graph_node_from_edge.add(v)
            node_list = list(set(node_list) & graph_node_from_edge)

           
            if len(node_list) != 0 :
                subgraph, subgraph_edge_index, mapping, _ = k_hop_subgraph(node_list, num_hops=args.num_hops, edge_index=cur_graph, relabel_nodes=True)
                vars(args)["subgraph"] = subgraph
                vars(args)["subgraph_edge_index"] = subgraph_edge_index
                vars(args)["mapping"] = mapping
            logger.info("number of increase nodes:{}, nodes after {} hop:{}, total nodes this year {}".format\
                        (len(node_list), args.num_hops, args.subgraph.size(), args.graph_size))
            vars(args)["node_list"] = np.asarray(node_list)
            vars(args)["warn_list"] = np.asarray(warn_list)


        

        if args.train:
            train(inputs, inputsP,inputsF, args)
        else:
            if args.auto_test:
                model= load_best_model(args)
                test_loader = DataLoader(COINDataset(inputs, "test"), batch_size=args.batch_size, shuffle=False, pin_memory=pin_memory, num_workers=n_work)
                test_model(model, args, test_loader, pin_memory=True)


    for i in [3, 6, 12]:
        for j in ['mae', 'rmse', 'mape']:
            info = ""
            for year in range(args.begin_year, args.end_year+1):
                if i in result:
                    if j in result[i]:
                        if year in result[i][j]:
                            info+="{:.2f}\t".format(result[i][j][year])
            logger.info("{}\t{}\t".format(i,j) + info)
    lis = []
    for i in [3, 6, 12]:
        for j in ['mae', 'rmse', 'mape']:
            info = ""
            count = 0
            averesult = 0
            for year in range(args.begin_year, args.end_year+1):
                if i in result:
                    if j in result[i]:
                        if year in result[i][j]:
                            averesult =  averesult+result[i][j][year]
                            count = count + 1
            averesult = averesult/count
            lis.append(str(averesult))
            info+="{:.2f}\t".format(averesult)
            logger.info("{}\t{}\t".format(i,j) + info)

    for year in range(args.begin_year, args.end_year+1):
        if year in result:
            info = "year\t{}\ttotal_time\t{}\taverage_time\t{}\tepoch\t{}".format(year, result[year]["total_time"], result[year]["average_time"], result[year]['epoch_num'])
            logger.info(info)
    lis = lis + [args.cl,args.tao,args.warn_ratio]


    excel_path = 'results_detail_' + os.path.basename(args.conf)  + '.xlsx'

    conf_name = os.path.basename(args.conf)

    # Create DataFrames for all three result types
    def create_result_df(result_dict):
        data = []
        for i in [3, 6, 12]:
            for j in ['mae', 'rmse', 'mape']:
                row = {'Metric': f'{i}_{j}'}
                for year in range(args.begin_year, args.end_year+1):
                    if i in result_dict and j in result_dict[i] and year in result_dict[i][j]:
                        row[f'Year_{year}'] = result_dict[i][j][year]
                # Calculate average
                values = [v for k, v in row.items() if k != 'Metric']
                if values:
                    row['Average'] = sum(values) / len(values)
                data.append(row)
        return pd.DataFrame(data)

    all_results = create_result_df(result)
    new_results = create_result_df(result_new) if 'result_new' in globals() else pd.DataFrame()
    old_results = create_result_df(result_old) if 'result_old' in globals() else pd.DataFrame()

    # Save to Excel
    try:
        # Try to load existing file
        book = load_workbook(excel_path)
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            writer.book = book
            # Delete sheet if it exists already
            if conf_name in book.sheetnames:
                del book[conf_name]
            # Write the data
            all_results.to_excel(writer, sheet_name=conf_name, index=False)
            if not new_results.empty:
                new_results.to_excel(writer, sheet_name=f'{conf_name}_new', index=False)
            if not old_results.empty:
                old_results.to_excel(writer, sheet_name=f'{conf_name}_old', index=False)
    except FileNotFoundError:
        # Create new file if it doesn't exist
        with pd.ExcelWriter(excel_path) as writer:
            all_results.to_excel(writer, sheet_name=conf_name, index=False)
            if not new_results.empty:
                new_results.to_excel(writer, sheet_name=f'{conf_name}_new', index=False)
            if not old_results.empty:
                old_results.to_excel(writer, sheet_name=f'{conf_name}_old', index=False)

    logger.info(f"Results saved to Excel file: {excel_path}")



    
    with open('results.csv', 'a', newline = '') as f:
        writer = csv.writer(f)
        writer.writerow(lis)
    time_stamp = np.array(args.time_stamp)
    np.save('time_stamp.npy', time_stamp)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(formatter_class = argparse.RawTextHelpFormatter)
    parser.add_argument("--conf", type = str, default = "conf/Traffic.json", help='Hum: configure for humidity, Tem: configure for temperature, Traffic: configure for traffic')

    parser.add_argument("--paral", type = int, default = 0)
    parser.add_argument("--gpuid", type = int, default = 0)
    parser.add_argument("--lr", type = float, default = 0.01)
    parser.add_argument("--epoch", type = int, default = 50)
    parser.add_argument("--inform_size",type = int, default= 0.1)

    parser.add_argument("--Using_RI_to_select", type = bool, default = False)
    parser.add_argument("--Using_unseen", type = bool, default = False)

    parser.add_argument("--inf", type = int, default = 1)
    parser.add_argument("--buffer_selection", type = bool, default = True)
    parser.add_argument("--start_selection", type = int, default = 45)
    parser.add_argument("--ratio", type = float, default = 0.5)

    parser.add_argument("--current_buffer_initiate", type = bool, default = True)
    parser.add_argument("--logname", type = str, default = "info")
    parser.add_argument("--tao", type = float, default = 0.1)
    parser.add_argument("--cl", type = float, default = 0.1)  
    parser.add_argument('--outputstyle', type = str, default = 'Existing',help ='All: output result on all nodes, Existing: output result on Existing nodes, New: output result on New nodes')
    parser.add_argument("--warn_ratio", type = float, default = 0.1)  
    parser.add_argument("--load_first_year", type = int, default = 0, help="0: training first year, 1: load from model path of first year")
    parser.add_argument("--first_year_model_path", type = str, default = "res/district3F11T17/TrafficStream2021-05-09-11:56:33.516033/2011/27.4437.pkl", help='specify a pretrained model root')
    args = parser.parse_args()
    init(args)
    seed_set(13)
    vars(args)['o_loss'] = [16.265, 15.4431, 15.1376, 17.3629, 13.4656, 14.6947]
    #vars(args)['o_loss'] = [16.265, 14.289, 16.0524, 18.3295, 16.7023, 14.2729]
    #vars(args)['o_loss'] = [16.265, 15.7817, 15.5147, 14.9328, 13.2777, 13.5057]
    device = torch.device("cuda:{}".format(args.gpuid)) if torch.cuda.is_available() and args.gpuid != -1 else "cpu"
    vars(args)["device"] = device
    vars(args)['time_stamp'] = []
    vars(args)['list'] = []
    main(args)